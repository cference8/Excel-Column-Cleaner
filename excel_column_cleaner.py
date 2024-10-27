import pandas as pd
import os
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json
from PIL import Image
import platform
import subprocess
from customtkinter import CTkImage
import sys
from random import randint
from pdf2image import convert_from_path, convert_from_bytes
import string
import PyPDF2
from io import BytesIO
import traceback
import threading

# Function to locate resource files, works for both PyInstaller executable and dev environment
def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Use the user's home directory or appdata folder for storing JSON
def get_user_data_directory():
    """Return a path to a writable directory for saving user data like the JSON file."""
    home_dir = os.path.expanduser("~")
    app_data_dir = os.path.join(home_dir, ".excel_cleaner")  # Hidden folder in the user's home
    os.makedirs(app_data_dir, exist_ok=True)  # Create the directory if it doesn't exist
    return app_data_dir

# Paths to your resources
logo_path = resource_path("scribe-logo-final.png")
icon_path = resource_path("scribe-icon.ico")
unchecked_columns_path = os.path.join(get_user_data_directory(), "unchecked_columns.json")  # Save JSON in writable directory

# Global variables to hold the DataFrame, input path, and list of saved files
df = None
input_path = None
saved_files = []  # To track the list of files saved during the session
unchecked_columns = []  # Initialize globally
show_all_columns = True  # Toggle between showing all columns and hiding unchecked

# Function to toggle the display of all columns or just checked ones
def toggle_show_hide_columns():
    global show_all_columns
    # Save the current state of checkboxes before toggling
    for column, var in checkbox_vars.items():
        if var.get():
            if column in unchecked_columns:
                unchecked_columns.remove(column)
        else:
            if column not in unchecked_columns:
                unchecked_columns.append(column)
    
    # Toggle the view state
    show_all_columns = not show_all_columns
    create_column_checkboxes(df.columns)

def create_column_checkboxes(columns):
    # Clear any previous checkboxes
    for widget in columns_frame.winfo_children():
        widget.destroy()

    # Create checkboxes for each column
    for column in columns:
        # Re-create checkbox variables based on previous states
        var = ctk.BooleanVar(value=True if column not in unchecked_columns else False)
        checkbox_vars[column] = var

        if show_all_columns or var.get():  # Show all columns or just checked ones
            checkbox = ctk.CTkCheckBox(columns_frame, text=column, variable=var)
            checkbox.pack(anchor="w", padx=10, pady=5)

# Function to set column width using openpyxl
def set_column_widths(output_path):
    try:
        wb = load_workbook(output_path)  # Removed keep_vba=True
        ws = wb.active

        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter  # Get column letter (e.g., 'A', 'B')
            column_name = column_cells[0].value  # Get the column header name

            if column_name == 'Custom Message':
                ws.column_dimensions[column_letter].width = 43  # Approx. 600 pixels
                continue  # Skip dynamic adjustment for this column

            elif column_name == 'Image Path':
                ws.column_dimensions[column_letter].width = 14  # Approx. 100 pixels
                continue  # Skip dynamic adjustment for this column

            # Dynamic adjustment for other columns
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value and isinstance(cell.value, str):
                        # Exclude hyperlink formulas from length calculation
                        if cell.value.startswith('=HYPERLINK'):
                            # Extract display text from the HYPERLINK formula
                            parts = cell.value.split(',', 1)
                            if len(parts) > 1:
                                display_text = parts[1].strip(')"')
                                max_length = max(max_length, len(display_text))
                        else:
                            max_length = max(max_length, len(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Add extra padding
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print("Column widths set successfully.")
    except Exception as e:
        print(f"Error setting column widths: {e}")

# Function to process a single file based on selected columns with color applied to "Custom Message" column when file contains "Nurture"
from tkinter import messagebox

def process_file(input_path, selected_columns, label):
    try:
        global df
        if df is None:
            df = pd.read_excel(input_path)

        # Ensure the "Salutation" column is included, even if not selected by the user
        if "Salutation" in df.columns and "Salutation" not in selected_columns:
            selected_columns.append("Salutation")

        # Add "Image Path" column if it exists in df
        if "Image Path" in df.columns and "Image Path" not in selected_columns:
            selected_columns.append("Image Path")

        # Validate selected columns against the columns in the DataFrame
        missing_columns = [col for col in selected_columns if col not in df.columns]
        if missing_columns:
            return f"Error: These columns are not in the file: {missing_columns}"

        # Select the specified columns that exist (including Salutation and Image Path)
        df_selected = df[selected_columns]

        # Generate the output path by appending "- QC_CLEAN" to the file name
        output_path = os.path.splitext(input_path)[0] + " - QC_CLEAN.xlsx"

        # Convert the 'Image Path' column to hyperlinks if it exists
        if 'Image Path' in df_selected.columns:
            for idx in df_selected.index:
                img_path = df_selected.at[idx, 'Image Path']
                if pd.notna(img_path):
                    # Replace backslashes with forward slashes for Excel hyperlink
                    img_path_excel = img_path.replace('\\', '/')
                    df_selected.at[idx, 'Image Path'] = f'=HYPERLINK("{img_path_excel}", "Image {idx + 1}")'

        # Save the dataframe to the output Excel file (initial save)
        df_selected.to_excel(output_path, index=False, engine='openpyxl')

        # Check if the filename contains the word "Nurture"
        if "Nurture" in os.path.basename(input_path):
            # Apply color formatting if the file contains "Nurture"
            apply_color_to_custom_message(output_path)

            # Apply color based on Salutation ID only if "Nurture" is in the filename
            apply_color_based_on_salutation(output_path)

        # Adjust column widths after applying colors
        set_column_widths(output_path)

        # Return the output file path
        return output_path

    except PermissionError:
        # Show a dialog informing the user to close the file in Excel
        messagebox.showerror("File Open in Excel", "File is open in Excel.\nPlease close the file opened in Excel before proceeding.")
        return "Error: File is open in Excel."

    except Exception as e:
        traceback.print_exc()
        return f"An error occurred with {input_path}: {str(e)}"

# Function to apply alternating colors to the "Custom Message" column whenever its content changes
def apply_color_to_custom_message(output_path):
    try:
        # Load the workbook and select the active worksheet
        wb = load_workbook(output_path)  # Removed keep_vba=True
        ws = wb.active

        # Define alternating colors
        colors = [
            PatternFill(start_color="EF9854", end_color="EF9854", fill_type="solid"),  # Yellow
            PatternFill(start_color="4C74A4", end_color="4C74A4", fill_type="solid")   # Magenta
        ]

        # Find the "Custom Message" column and apply color only to cells in this column when the value changes
        custom_message_col = None
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == "Custom Message":
                custom_message_col = col[0].column_letter
                break

        if custom_message_col:
            previous_value = None
            current_color_idx = 0
            # Apply color to "Custom Message" column cells based on changes in the value
            for row_idx in range(2, ws.max_row + 1):  # Start at row 2 to skip header
                cell_value = ws[f"{custom_message_col}{row_idx}"].value
                if cell_value != previous_value:
                    current_color_idx = (current_color_idx + 1) % len(colors)  # Alternate colors
                previous_value = cell_value
                # Apply color to the cell in the "Custom Message" column
                ws[f"{custom_message_col}{row_idx}"].fill = colors[current_color_idx]

        # Save the workbook with the color formatting
        wb.save(output_path)
        print("Custom Message column colored successfully.")
    except Exception as e:
        print(f"Error applying color to Custom Message: {e}")

# Function to generate unique colors for each Salutation ID
def generate_unique_color():
    """Generate a visually distinct light color."""
    # Generate light colors by keeping the RGB values above a certain threshold
    r = randint(150, 255)  # Higher values to ensure lighter colors
    g = randint(150, 255)
    b = randint(150, 255)

    # Create the color in hexadecimal format
    color_hex = f'{r:02X}{g:02X}{b:02X}'
    
    # Return a PatternFill with the light color
    return PatternFill(
        start_color=color_hex,
        end_color=color_hex,
        fill_type="solid"
    )

# Function to apply unique colors based on "Salutation" ID to selected columns
def apply_color_based_on_salutation(output_path):
    try:
        # Load the workbook and select the active worksheet
        wb = load_workbook(output_path)  # Removed keep_vba=True
        ws = wb.active

        # Generate a unique color for each "Salutation" ID
        salutation_colors = {}
        salutation_col_letter = None
        selected_columns_letters = []

        # Find the "Salutation" column
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == "Salutation":
                salutation_col_letter = col[0].column_letter
                break

        if not salutation_col_letter:
            raise ValueError("Salutation column not found.")

        # Find selected columns (excluding "Custom Message")
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value != "Custom Message":
                selected_columns_letters.append(col[0].column_letter)

        # Loop through the rows to apply unique colors based on the "Salutation" ID
        for row_idx in range(2, ws.max_row + 1):  # Start at row 2 to skip header
            salutation_value = ws[f"{salutation_col_letter}{row_idx}"].value

            if salutation_value not in salutation_colors:
                # Assign a unique light color to each new Salutation ID
                salutation_colors[salutation_value] = generate_unique_color()

            # Apply the assigned color to all selected columns (excluding "Custom Message")
            for col_letter in selected_columns_letters:
                ws[f"{col_letter}{row_idx}"].fill = salutation_colors[salutation_value]

        # Save the workbook with the applied color formatting
        wb.save(output_path)
        print("Salutation-based coloring applied successfully.")
    except Exception as e:
        print(f"Error applying color based on Salutation: {e}")

# Function to save unchecked columns to a file (updated to keep unique values)
def save_unchecked_columns(unchecked_columns, file_path=unchecked_columns_path):
    try:
        existing_columns = load_unchecked_columns(file_path)
        # Combine existing and new unchecked columns and keep unique values
        all_unchecked_columns = list(set(existing_columns) | set(unchecked_columns))
        
        with open(file_path, "w") as file:
            json.dump(all_unchecked_columns, file)
        print("Unchecked columns saved successfully.")
    except Exception as e:
        print(f"Error saving unchecked columns: {e}")

# Function to load unchecked columns from a file (updated to load unique values)
def load_unchecked_columns(file_path=unchecked_columns_path):
    try:
        if os.path.exists(file_path):
            with open(file_path, "r") as file:
                return json.load(file)
        return []
    except Exception as e:
        print(f"Error loading unchecked columns: {e}")
        return []

# Function to clear the uploaded file and reset the column list in the GUI
def clear_gui_after_processing():
    global df, input_path, checkbox_vars
    df = None
    input_path = None

    # Clear the checkboxes in the GUI
    for widget in columns_frame.winfo_children():
        widget.destroy()

    # Clear the dictionary that tracks the state of checkboxes for each column
    checkbox_vars.clear()

    # Disable the process button
    process_button.configure(state="disabled")
    print("GUI reset after processing.")

# Function to update the saved files label with the current list
def update_saved_files_label():
    saved_files_text = "\n".join(saved_files) if saved_files else "No files saved yet."
    saved_files_label.configure(text=f"Files saved in this session:\n{saved_files_text}\n _________________________")
    print("Saved files label updated.")

# Function to open file explorer in the directory of the output file
def open_file_explorer(output_path):
    try:
        directory = os.path.dirname(output_path)
        if platform.system() == "Windows":
            os.startfile(directory)
        elif platform.system() == "Darwin":  # macOS
            subprocess.Popen(["open", directory])
        else:  # Linux
            subprocess.Popen(["xdg-open", directory])
        print(f"Opened file explorer at {directory}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file explorer: {str(e)}")
        print(f"Error opening file explorer: {e}")

# Function to get selected columns and process the file
def get_selected_columns_and_process():
    selected_columns = [col for col, var in checkbox_vars.items() if var.get()]
    unchecked_columns_local = [col for col, var in checkbox_vars.items() if not var.get()]

    if selected_columns:
        global input_path
        output_path = process_file(input_path, selected_columns, output_label)

        if "Error" in output_path:
            output_label.configure(text=output_path)
        else:
            # Display the file save location in the label
            output_label.configure(text=f"File saved successfully at: {output_path}")

            # Add the saved file to the running list
            saved_files.append(output_path)

            # Update the saved files label
            update_saved_files_label()

            # Open the file explorer in the directory of the saved file
            open_file_explorer(output_path)
            
            # Clear the uploaded file and reset the column list in the GUI after processing
            clear_gui_after_processing()

        # Save unchecked columns (now it will preserve unique values)
        save_unchecked_columns(unchecked_columns_local)

    else:
        messagebox.showwarning("No columns selected", "Please select at least one column.")
        print("No columns selected by the user.")

# Function to upload the file and show columns for selection
def upload_file_and_show_columns():
    global input_path, unchecked_columns
    input_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if input_path:
        try:
            global df
            df = pd.read_excel(input_path)

            # Update Excel file label
            excel_file_label.configure(text=f"Excel file uploaded: {os.path.basename(input_path)}")

            # Load the previously unchecked columns
            unchecked_columns = load_unchecked_columns()

            # Create the checkboxes for the columns
            create_column_checkboxes(df.columns)

            # Enable the "Process File" button after file is loaded
            process_button.configure(state="normal")

            # Update the canvas scroll region
            canvas.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))
            print("Excel file uploaded and checkboxes created.")

            # Check if "Nurture" is in the Excel file name (case-insensitive)
            if "nurture" in os.path.basename(input_path).lower():
                # Prompt the user to select a PDF file
                pdf_path = filedialog.askopenfilename(
                    title="Select Nurture Single Card Design PDF",
                    filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
                )

                if pdf_path:
                    # Update PDF file label
                    pdf_file_label.configure(text=f"PDF file uploaded: {os.path.basename(pdf_path)}")
                    
                    base_image_dir = r'G:\Shared drives\Scribe Workspace\Scribe Master Folder\Scribe Nurture\Scribe Intellicut Design Files\Nurture Image Previews'
                    # Sanitize the PDF file name to create a valid directory name
                    pdf_file_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    pdf_file_name = sanitize_filename(pdf_file_name)
                    images_dir = os.path.join(base_image_dir, pdf_file_name)

                    # Check if the directory already exists
                    if os.path.exists(images_dir):
                        # Prompt the user with a Yes/No dialog
                        regenerate = messagebox.askyesno(
                            "Images Already Exist",
                            "That PDF file has already been converted to images. Would you like to regenerate the images?"
                        )
                        if not regenerate:
                            # If user selects 'No', load existing images and update DataFrame
                            existing_image_paths = [os.path.join(images_dir, f) for f in sorted(os.listdir(images_dir))]
                            df['Image Path'] = existing_image_paths[:500]  # Use only first 500 images if more are present
                            messagebox.showinfo("Images Loaded", "Existing images loaded successfully.")
                            return

                    # Create the directory for images if regeneration is needed
                    os.makedirs(images_dir, exist_ok=True)

                    # Show the progress bar and percentage label
                    progress_label.configure(text="Converting Single Card Design PDF to Images:")
                    progress_label.grid()  # Make the progress label visible
                    progress_value_label.configure(text="0%")
                    progress_value_label.grid()  # Make the percentage label visible
                    progress_bar.set(0)
                    progress_bar.grid()  # Make the progress bar visible

                    # Disable the "Process File" button during processing
                    process_button.configure(state="disabled")

                    # Start the PDF processing in a separate thread
                    thread = threading.Thread(target=lambda: handle_pdf_processing(pdf_path, images_dir))
                    thread.start()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
            print(f"Error loading Excel file: {e}")

def handle_pdf_processing(pdf_path, images_dir):
    try:
        image_paths = process_pdf(pdf_path, images_dir)
        if image_paths:
            num_images = len(image_paths)
            if num_images != 500:
                messagebox.showwarning("Warning", f"The PDF should produce exactly 500 images from odd pages. Found {num_images} images.")
                print(f"Expected 500 images, but found {num_images}.")
                return

            # Check if DataFrame has at least 500 rows
            if len(df) < 500:
                messagebox.showwarning("Warning", f"The Excel file should have at least 500 rows. Found {len(df)} rows.")
                print(f"DataFrame has {len(df)} rows, which is less than 500.")
                return

            # Add image paths to the DataFrame
            df['Image Path'] = image_paths[:500]  # Ensure only first 500 are used

            messagebox.showinfo("Success", "PDF images extracted and image paths added to the Excel data.")
            print("PDF processing completed successfully.")
        else:
            messagebox.showerror("Error", "Failed to process the PDF.")
            print("PDF processing failed.")
    except Exception as e:
        print(f"Error during PDF processing: {e}")
        messagebox.showerror("Error", "An unexpected error occurred during PDF processing.")
    finally:
        # Re-enable the "Process File" button after processing completes
        root.after(0, lambda: process_button.configure(state="normal"))

# PDF TO IMAGE Functions
def extract_images_from_pdf(pdf_path, output_folder, progress_callback=None):
    import PyPDF2
    from pdf2image import convert_from_bytes
    from io import BytesIO

    image_paths = []
    idx = 0

    # Open the PDF file and keep it open during processing
    with open(pdf_path, 'rb') as f:
        pdf_reader = PyPDF2.PdfReader(f)
        total_pages = len(pdf_reader.pages)

        # Calculate the total number of odd pages
        total_odd_pages = (total_pages + 1) // 2

        # Process odd-numbered pages (zero-based indexing: 0, 2, 4, ...)
        for current, page_number in enumerate(range(0, total_pages, 2), start=1):
            try:
                pdf_writer = PyPDF2.PdfWriter()
                pdf_writer.add_page(pdf_reader.pages[page_number])
                
                # Write the single page to a BytesIO object
                pdf_bytes = BytesIO()
                pdf_writer.write(pdf_bytes)
                pdf_bytes.seek(0)
                
                # Convert the single-page PDF to an image
                images = convert_from_bytes(pdf_bytes.read())
                
                for image in images:
                    idx += 1
                    image_filename = os.path.join(output_folder, f"image_{idx}.jpg")
                    image.save(image_filename, 'JPEG')
                    image_paths.append(image_filename)
            
                # Update progress
                if progress_callback:
                    progress = current / total_odd_pages  # Progress as a float between 0 and 1
                    progress_callback(progress)
            except Exception as e:
                print(f"Error processing page {page_number + 1}: {e}")
                continue  # Skip to the next page in case of an error

    return image_paths

def process_pdf(pdf_path, images_dir):
    def progress_callback(progress):
        # Schedule the progress bar update in the main thread
        root.after(0, update_progress_bar, progress)

    try:
        image_paths = extract_images_from_pdf(pdf_path, images_dir, progress_callback)
        return image_paths
    except Exception as e:
        print(f"Error in thread: {e}")
        return None

def update_progress_bar(progress):
    progress_bar.set(progress)  # Update the progress bar value
    percentage = int(progress * 100)
    progress_value_label.configure(text=f"{percentage}%")  # Update the percentage label
    if progress >= 1.0:
        progress_bar.grid_remove()  # Hide the progress bar when done
        progress_label.configure(text="PDF Processing Complete:")
        progress_value_label.configure(text="100%")  # Ensure it's set to 100% at completion
        progress_value_label.grid_remove()  # Hide the percentage label after completion

def sanitize_filename(filename):
    valid_chars = f"-_.() {string.ascii_letters}{string.digits}"
    sanitized = ''.join(c for c in filename if c in valid_chars)
    return sanitized

# Function to handle mouse wheel scrolling
def on_mousewheel(event):
    _on_mousewheel(event)

# Internal function to handle cross-platform mouse wheel scrolling
def _on_mousewheel(event):
    if platform.system() == 'Windows':
        delta = int(-1 * (event.delta / 120))
    elif platform.system() == 'Darwin':  # macOS
        delta = int(-1 * (event.delta))
    else:  # Linux
        delta = int(-1 * (event.delta / 120))
    canvas.yview_scroll(delta, "units")

# Setting up the CustomTkinter GUI window
ctk.set_appearance_mode("light")  # Options: "light", "dark", "system"
ctk.set_default_color_theme("blue")  # Options: "blue", "dark-blue", "green"

# Create the main window
root = ctk.CTk()
root.geometry("600x600")  # Increased size for better visibility
root.title("Excel Column Cleaner")

# Load the custom icon in .ico format (replace the default CustomTkinter icon)
icon_path = resource_path('scribe-icon.ico')
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)
else:
    print(f"Icon file not found at {icon_path}")

# Global variable to store checkboxes and their states
checkbox_vars = {}

# Adding the logo at the top of the window using CTkImage for HighDPI displays
if os.path.exists(logo_path):
    img = Image.open(logo_path)
    img_ctk = CTkImage(light_image=img, size=(258, 100))
    
    logo_label = ctk.CTkLabel(root, image=img_ctk, text="")
    logo_label.grid(row=0, column=0, columnspan=2, pady=10)
else:
    print(f"Logo file not found at {logo_path}")
    logo_label = ctk.CTkLabel(root, text="Excel Column Cleaner", font=("Arial", 24, "bold"))
    logo_label.grid(row=0, column=0, columnspan=2, pady=10)

# Input file selection label
header_label = ctk.CTkLabel(root, text="Select the Excel file | columns to keep:", font=("Arial", 16, "bold"), wraplength=300)
header_label.grid(row=1, column=0, pady=10, padx=10, sticky="w")

# Additional labels to display uploaded file names
excel_file_label = ctk.CTkLabel(root, text="No Excel file uploaded", font=("Arial", 12), justify="left")
excel_file_label.grid(row=2, column=0, pady=5, padx=10, sticky="w")

pdf_file_label = ctk.CTkLabel(root, text="No PDF file uploaded", font=("Arial", 12), justify="left")
pdf_file_label.grid(row=3, column=0, pady=5, padx=10, sticky="w")
pdf_file_label.grid_remove()

# Browse button to select files
browse_button = ctk.CTkButton(root, text="Browse File", font=("Arial", 14), command=upload_file_and_show_columns)
browse_button.grid(row=4, column=0, pady=10, padx=10, sticky="ew")

# Progress bar label
progress_label = ctk.CTkLabel(root, text="Converting Single Card Design PDF to Images:", font=("Arial", 12), justify="left")
progress_label.grid(row=5, column=0, pady=(10, 0), padx=10, sticky="w")
progress_label.grid_remove()  # Hide initially

# Progress bar widget
progress_bar = ctk.CTkProgressBar(root, width=250)
progress_bar.grid(row=6, column=0, pady=(0, 10), padx=10, sticky="w")
progress_bar.set(0)  # Initialize to 0%
progress_bar.grid_remove()  # Hide initially

# Progress percentage label
progress_value_label = ctk.CTkLabel(root, text="0%", font=("Arial", 12), justify="left")
progress_value_label.grid(row=6, column=1, pady=(0, 10), padx=(0,50), sticky="w")
progress_value_label.grid_remove()  # Hide initially

# Process button to generate the new file based on selected columns
process_button = ctk.CTkButton(root, text="Create Clean Excel File", font=("Arial", 14), command=get_selected_columns_and_process, state="disabled")
process_button.grid(row=7, column=0, pady=10, padx=10, sticky="ew")

# Show/Hide Unchecked Columns button
toggle_button = ctk.CTkButton(root, text="Show/Hide Unchecked Columns", font=("Arial", 14), command=toggle_show_hide_columns)
toggle_button.grid(row=8, column=0, pady=10, padx=10, sticky="ew")

# Output label to display the file save paths or errors
output_label = ctk.CTkLabel(root, text="", wraplength=250, font=("Arial", 12), justify="left")
output_label.grid(row=9, column=0, pady=10, padx=10, sticky="w")

# Saved files label to display a running list of saved files
saved_files_label = ctk.CTkLabel(root, text="Files saved in this session:\nNo files saved yet.", wraplength=250, font=("Arial", 12), justify="left")
saved_files_label.grid(row=10, column=0, pady=10, padx=10, sticky="w")

# Frame for column selection checkboxes with scrollable canvas
columns_frame_container = ctk.CTkFrame(root)
columns_frame_container.grid(row=1, column=1, rowspan=8, padx=(30,5), pady=10, sticky="nsw")  # Adjusted rowspan

# Set the background color for the canvas
canvas = ctk.CTkCanvas(columns_frame_container, bg='#f0f0f0')  # Light grey background
canvas.pack(side="left", fill="both", expand=True)

# Create an inner frame on the canvas with the same background color
columns_frame = ctk.CTkFrame(canvas, bg_color='#f0f0f0')  # Light grey background
canvas_frame = canvas.create_window((0, 0), window=columns_frame, anchor="nw")

# Bind mouse wheel scrolling to the canvas
canvas.bind_all("<MouseWheel>", _on_mousewheel)

# Adjust the grid weights for resizing
root.grid_rowconfigure(0, weight=0)  # Logo row
root.grid_rowconfigure(1, weight=1)  # Column selection row
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)  
root.grid_rowconfigure(4, weight=0)  # Progress label row
root.grid_rowconfigure(5, weight=0)  # Progress bar row
root.grid_rowconfigure(6, weight=1)
root.grid_rowconfigure(7, weight=1)
root.grid_rowconfigure(8, weight=1)
root.grid_rowconfigure(9, weight=1)
root.grid_rowconfigure(10, weight=1)
root.grid_columnconfigure(0, weight=0)  # Left side
root.grid_columnconfigure(1, weight=1)  # Right side

# Run the CustomTkinter event loop
root.mainloop()
