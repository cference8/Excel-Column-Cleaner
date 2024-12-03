import pandas as pd
import os
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json
from PIL import Image
import platform
import subprocess
from customtkinter import CTkImage
import sys
from random import randint

# Function to locate resource files, works for both PyInstaller executable and dev environment
def resource_path(relative_path):
    """Get the absolute path to a resource, works for dev and for PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Use the user's home directory or appdata folder for storing JSON
def get_user_data_directory():
    """Return a path to a writable directory for saving user data like the JSON file."""
    home_dir = os.path.expanduser("~")
    app_data_dir = os.path.join(home_dir, ".excel_cleaner")  # Hidden folder in the user's home
    os.makedirs(app_data_dir, exist_ok=True)  # Create the directory if it doesn't exist
    return app_data_dir

unchecked_columns_path = os.path.join(get_user_data_directory(), "unchecked_columns.json")  # Save JSON in writable directory

# Global variables to hold the DataFrame, input path, and list of saved files
df = None
input_path = None
saved_files = []  # To track the list of files saved during the session
unchecked_columns = []  # Initialize globally
show_all_columns = True  # Toggle between showing all columns and hiding unchecked

# Function to toggle the display of all columns or just checked ones
def toggle_show_hide_columns():
    global show_all_columns
    # Save the current state of checkboxes before toggling
    for column, var in checkbox_vars.items():
        if var.get():
            if column in unchecked_columns:
                unchecked_columns.remove(column)
        else:
            if column not in unchecked_columns:
                unchecked_columns.append(column)

    # Toggle the view state
    show_all_columns = not show_all_columns
    create_column_checkboxes(df.columns)

def create_column_checkboxes(columns):
    # Clear any previous checkboxes
    for widget in columns_frame.winfo_children():
        widget.destroy()

    # Create checkboxes for each column
    for column in columns:
        # Re-create checkbox variables based on previous states
        var = ctk.BooleanVar(value=True if column not in unchecked_columns else False)
        checkbox_vars[column] = var

        if show_all_columns or var.get():  # Show all columns or just checked ones
            checkbox = ctk.CTkCheckBox(columns_frame, text=column, variable=var)
            checkbox.pack(anchor="w", padx=10, pady=5)

# Function to set column width using openpyxl
def set_column_widths(output_path):
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        max_column_width = 30  # Maximum column width
        padding = 1  # Reduced padding

        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            column_name = column_cells[0].value

            if column_name == 'Custom Message':
                ws.column_dimensions[column_letter].width = 50  # Adjusted width
                continue

            elif column_name == 'Outer Design File':
                ws.column_dimensions[column_letter].width = 15  # Adjusted width
                continue

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            adjusted_width = min(max_length + padding, max_column_width)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print("Column widths set successfully.")
    except Exception as e:
        print(f"Error setting column widths: {e}")

# Function to process a single file based on selected columns with color applied to "Custom Message" column when file contains "Nurture"
def process_file(input_path, selected_columns, label):
    try:
        global df
        if df is None:
            df = pd.read_excel(input_path)

        # Ensure the "Salutation" column is included, even if not selected by the user
        if "Salutation" in df.columns and "Salutation" not in selected_columns:
            selected_columns.append("Salutation")

        # Ensure "Outer Design File" column is included if it exists
        if "Outer Design File" in df.columns and "Outer Design File" not in selected_columns:
            selected_columns.append("Outer Design File")

        # Validate selected columns against the columns in the DataFrame
        missing_columns = [col for col in selected_columns if col not in df.columns]
        if missing_columns:
            raise Exception(f"Error: These columns are not in the file: {missing_columns}")

        # Select the specified columns that exist (including Salutation and Outer Design File)
        df_selected = df[selected_columns]

        # Check if the filename contains the word "Nurture"
        if "nurture" in os.path.basename(input_path).lower():
            # Process the "Outer Design File" column
            if "Outer Design File" in df_selected.columns:
                # Wrap URLs in the "Outer Design File" column into Excel HYPERLINK formulas
                for idx, url in df_selected["Outer Design File"].items():  # Changed iteritems() to items()
                    if pd.notna(url):
                        row_num = idx + 2  # Excel row number (starting from 2)
                        hyperlink_formula = f'=HYPERLINK("{url}", "Row {row_num} Image")'
                        df_selected.at[idx, "Outer Design File"] = hyperlink_formula

        # Generate the output path by appending "- QC_CLEAN" to the file name
        output_path = os.path.splitext(input_path)[0] + " - QC_CLEAN.xlsx"

        # Save the dataframe to the output Excel file (initial save)
        df_selected.to_excel(output_path, index=False)

        # Check if the filename contains the word "Nurture"
        if "nurture" in os.path.basename(input_path).lower():
            # Apply color formatting if the file contains "Nurture"
            apply_color_to_custom_message(output_path)

            # Apply color based on Salutation ID only if "Nurture" is in the filename
            apply_color_based_on_salutation(output_path)

        # Adjust column widths after applying colors
        set_column_widths(output_path)

        # Return the output file path
        return output_path

    except Exception as e:
        print(f"Exception in process_file: {e}")
        raise Exception(f"An error occurred with {input_path}: {str(e)}")

# Function to apply alternating colors to the "Custom Message" column whenever its content changes
def apply_color_to_custom_message(output_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(output_path)
    ws = wb.active

    # Define alternating colors
    colors = [
        PatternFill(start_color="ef9854", end_color="ef9854", fill_type="solid"),  # Yellow
        PatternFill(start_color="4c74a4", end_color="4c74a4", fill_type="solid")   # Magenta
    ]

    # Find the "Custom Message" column and apply color only to cells in this column when the value changes
    custom_message_col = None
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        if col[0].value == "Custom Message":
            custom_message_col = col[0].column_letter
            break

    if custom_message_col:
        previous_value = None
        current_color_idx = 0
        # Apply color to "Custom Message" column cells based on changes in the value
        for row_idx in range(2, ws.max_row + 1):  # Start at row 2 to skip header
            cell_value = ws[f"{custom_message_col}{row_idx}"].value
            if cell_value != previous_value:
                current_color_idx = (current_color_idx + 1) % len(colors)  # Alternate colors
            previous_value = cell_value
            # Apply color to the cell in the "Custom Message" column
            ws[f"{custom_message_col}{row_idx}"].fill = colors[current_color_idx]

    # Save the workbook with the color formatting
    wb.save(output_path)

# Function to generate unique colors for each Salutation ID
def generate_unique_color():
    """Generate a visually distinct light color."""
    # Generate light colors by keeping the RGB values above a certain threshold
    r = randint(150, 255)  # Higher values to ensure lighter colors
    g = randint(150, 255)
    b = randint(150, 255)

    # Create the color in hexadecimal format
    color_hex = f'{r:02X}{g:02X}{b:02X}'

    # Return a PatternFill with the light color
    return PatternFill(
        start_color=color_hex,
        end_color=color_hex,
        fill_type="solid"
    )


# Function to apply unique colors based on "Salutation" ID to selected columns
def apply_color_based_on_salutation(output_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(output_path)
    ws = wb.active

    # Generate a unique color for each "Salutation" ID
    salutation_colors = {}
    salutation_col_letter = None
    selected_columns_letters = []

    # Find the "Salutation" column
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        if col[0].value == "Salutation":
            salutation_col_letter = col[0].column_letter
            break

    if not salutation_col_letter:
        raise ValueError("Salutation column not found.")

    # Find selected columns (excluding "Custom Message")
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        if col[0].value != "Custom Message":
            selected_columns_letters.append(col[0].column_letter)

    # Loop through the rows to apply unique colors based on the "Salutation" ID
    for row_idx in range(2, ws.max_row + 1):  # Start at row 2 to skip header
        salutation_value = ws[f"{salutation_col_letter}{row_idx}"].value

        if salutation_value not in salutation_colors:
            # Assign a unique light color to each new Salutation ID
            salutation_colors[salutation_value] = generate_unique_color()

        # Apply the assigned color to all selected columns (excluding "Custom Message")
        for col_letter in selected_columns_letters:
            ws[f"{col_letter}{row_idx}"].fill = salutation_colors[salutation_value]

    # Save the workbook with the applied color formatting
    wb.save(output_path)

# Function to save unchecked columns to a file (updated to keep unique values)
def save_unchecked_columns(unchecked_columns, file_path=unchecked_columns_path):
    existing_columns = load_unchecked_columns(file_path)
    # Combine existing and new unchecked columns and keep unique values
    all_unchecked_columns = list(set(existing_columns) | set(unchecked_columns))

    with open(file_path, "w") as file:
        json.dump(all_unchecked_columns, file)

# Function to load unchecked columns from a file (updated to load unique values)
def load_unchecked_columns(file_path=unchecked_columns_path):
    if os.path.exists(file_path):
        with open(file_path, "r") as file:
            return json.load(file)
    return []

# Function to clear the uploaded file and reset the column list in the GUI
def clear_gui_after_processing():
    global df, input_path, checkbox_vars
    df = None
    input_path = None

    # Clear the checkboxes in the GUI
    for widget in columns_frame.winfo_children():
        widget.destroy()

    # Clear the dictionary that tracks the state of checkboxes for each column
    checkbox_vars.clear()

    # Disable the process button
    process_button.configure(state="disabled")

# Function to update the saved files label with the current list
def update_saved_files_label():
    saved_files_text = "\n".join(saved_files) if saved_files else "No files saved yet."
    saved_files_label.configure(text=f"Files saved in this session:\n{saved_files_text}\n _________________________")

# Function to open file explorer in the directory of the output file
def open_file_explorer(output_path):
    try:
        directory = os.path.dirname(output_path)
        if platform.system() == "Windows":
            os.startfile(directory)
        elif platform.system() == "Darwin":  # macOS
            subprocess.Popen(["open", directory])
        else:  # Linux
            subprocess.Popen(["xdg-open", directory])
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file explorer: {str(e)}")

def get_selected_columns_and_process():
    selected_columns = [col for col, var in checkbox_vars.items() if var.get()]
    unchecked_columns_local = [col for col, var in checkbox_vars.items() if not var.get()]

    if selected_columns:
        global input_path
        try:
            output_path = process_file(input_path, selected_columns, output_label)
            # Display the file save location in the label
            output_label.configure(text=f"File saved successfully at: {output_path}")

            # Add the saved file to the running list
            saved_files.append(output_path)

            # Update the saved files label
            update_saved_files_label()

            # Open the file explorer in the directory of the saved file
            open_file_explorer(output_path)
        except Exception as e:
            # Display the error message
            output_label.configure(text=str(e))
            print(f"Error during processing: {e}")
            messagebox.showerror("Error", str(e))
        finally:
            # Save unchecked columns (now it will preserve unique values)
            save_unchecked_columns(unchecked_columns_local)

            # Clear the uploaded file and reset the column list in the GUI after processing
            clear_gui_after_processing()
    else:
        messagebox.showwarning("No columns selected", "Please select at least one column.")

# Function to upload the file and show columns for selection
def upload_file_and_show_columns():
    global input_path, unchecked_columns
    input_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if input_path:
        global df
        df = pd.read_excel(input_path)

        # Update the file_name_label with the selected file name
        file_name = os.path.basename(input_path)
        file_name_label.configure(text=f"Selected file: {file_name}")

        # Load the previously unchecked columns
        unchecked_columns = load_unchecked_columns()

        # Create the checkboxes for the columns
        create_column_checkboxes(df.columns)

        # Enable the "Process File" button after file is loaded
        process_button.configure(state="normal")

        # Update the canvas scroll region
        canvas.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

# Function to handle mouse wheel scrolling
def on_mousewheel(event):
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

# Setting up the CustomTkinter GUI window
ctk.set_appearance_mode("dark")  # Options: "light", "dark", "system"
ctk.set_default_color_theme("blue")  # Options: "blue", "dark-blue", "green"

# Create the main window
root = ctk.CTk()
root.geometry("600x600")
root.title("Excel Optimizer")

# Load the custom icon in .ico format (replace the default CustomTkinter icon)
icon_path = resource_path('scribe-icon.ico')
root.iconbitmap(icon_path)

# Global variable to store checkboxes and their states
checkbox_vars = {}

# Paths to your resources (adjusted to handle .webp and .png fallback)
logo_path_webp = resource_path("scribe-logo-final.webp")
logo_path_png = resource_path("scribe-logo-final.png")

# Attempt to load the .webp image; fallback to .png if .webp is unavailable
try:
    if os.path.exists(logo_path_webp):
        img = Image.open(logo_path_webp)
    elif os.path.exists(logo_path_png):
        img = Image.open(logo_path_png)
    else:
        raise FileNotFoundError("Neither .webp nor .png logo file found.")
except Exception as e:
    print(f"Error loading logo: {e}")
    img = None

if img:
    img_ctk = CTkImage(light_image=img, size=(258, 100))
    logo_label = ctk.CTkLabel(root, image=img_ctk, text="")
else:
    # If no image is loaded, use a default text placeholder
    logo_label = ctk.CTkLabel(root, text="Logo Missing", font=("Arial", 16))

# Add the logo label to the GUI
logo_label.grid(row=0, column=0, columnspan=2, pady=10)

# Input file selection label
header_label = ctk.CTkLabel(root, text="Select Excel file & columns to keep:", font=("Arial", 16, "bold"), wraplength=300)
header_label.grid(row=1, column=0, pady=10, padx=10)

# Label to display the uploaded file name
file_name_label = ctk.CTkLabel(root, text="No file selected", font=("Arial", 12))
file_name_label.grid(row=2, column=0, pady=5, padx=10, sticky="w")

# Browse button to select files
browse_button = ctk.CTkButton(root, text="Select Excel File", font=("Arial", 14), command=upload_file_and_show_columns)
browse_button.grid(row=3, column=0, pady=10, padx=10, sticky="ew")

# Process button to generate the new file based on selected columns
process_button = ctk.CTkButton(root, text="Create clean Excel File", font=("Arial", 14), command=get_selected_columns_and_process, state="disabled")
process_button.grid(row=4, column=0, pady=10, padx=10, sticky="ew")

# Show/Hide Unchecked Columns button to toggle between showing all columns and hiding unchecked
toggle_button = ctk.CTkButton(root, text="Show/Hide Unchecked Columns", font=("Arial", 14), command=toggle_show_hide_columns)
toggle_button.grid(row=5, column=0, pady=10, padx=10, sticky="ew")

# Output label to display the file save paths or errors
output_label = ctk.CTkLabel(root, text="", wraplength=250, font=("Arial", 12), justify="left")
output_label.grid(row=6, column=0, pady=20)

# Saved files label to display a running list of saved files
saved_files_label = ctk.CTkLabel(root, text="Files saved in this session:\nNo files saved yet.", wraplength=250, font=("Arial", 12), justify="left")
saved_files_label.grid(row=7, column=0, pady=20)

# Frame for column selection checkboxes with scrollable canvas
columns_frame_container = ctk.CTkFrame(root)
columns_frame_container.grid(row=1, column=1, rowspan=6, padx=20, pady=10, sticky="nsew")  # Adjust position to the right side

# Set the background color for the canvas
canvas = ctk.CTkCanvas(columns_frame_container, width=260)  # Grey background
canvas.pack(side="left", fill="both", expand=True)

scrollbar = ctk.CTkScrollbar(columns_frame_container, orientation="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")

canvas.configure(yscrollcommand=scrollbar.set)

# Create an inner frame on the canvas with the same background color
columns_frame = ctk.CTkFrame(canvas, bg_color='white')  # Grey background
canvas_frame = canvas.create_window((5, 5), window=columns_frame, anchor="nw")

# Bind mouse wheel scrolling to the canvas
canvas.bind_all("<MouseWheel>", on_mousewheel)

# Adjust the grid weights for resizing
root.grid_rowconfigure(0, weight=0)  # Logo row
root.grid_rowconfigure(1, weight=0)  # Header label row
root.grid_rowconfigure(2, weight=0)  # File name label row
root.grid_rowconfigure(3, weight=0)  # Browse button row
root.grid_rowconfigure(4, weight=0)  # Process button row
root.grid_rowconfigure(5, weight=0)  # Toggle button row
root.grid_rowconfigure(6, weight=1)  # Output label row
root.grid_rowconfigure(7, weight=1)  # Saved files label row

root.grid_columnconfigure(0, weight=1)  # Left side
root.grid_columnconfigure(1, weight=3)  # Right side


# Run the CustomTkinter event loop
root.mainloop()